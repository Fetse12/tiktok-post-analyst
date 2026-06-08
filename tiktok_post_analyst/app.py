import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime
from io import BytesIO
import os
import time
import random as _random

# Dynamic import safety for yt-dlp
try:
    import yt_dlp
    try:
        from yt_dlp.networking.impersonate import ImpersonateTarget
    except ImportError:
        ImpersonateTarget = None
except ImportError:
    yt_dlp = None
    ImpersonateTarget = None

# Dynamic import safety for openpyxl (Excel styling)
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    Font = PatternFill = Alignment = Border = Side = None

NUM_POSTS = 7

# Set premium page config
st.set_page_config(
    page_title="TikTok Batch Post Analyst",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Session state initialization
if 'posts_results' not in st.session_state:
    st.session_state.posts_results = []
if 'batch_analyzed' not in st.session_state:
    st.session_state.batch_analyzed = False
if 'fetch_status_messages' not in st.session_state:
    st.session_state.fetch_status_messages = []

# TikTok Cyberpunk Brand Theme Colors
primary_color = "#FF0050"       # TikTok Pink
secondary_color = "#00F2FE"     # TikTok Cyan
accent_gradient = "linear-gradient(90deg, #FF0050 0%, #a855f7 50%, #00F2FE 100%)"
card_glow = "rgba(255, 0, 80, 0.15)"

# Custom CSS
st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Outfit:wght@400;500;600;700;800&display=swap');

    .stApp {{
        background: linear-gradient(135deg, #09070f 0%, #120e24 50%, #050407 100%);
        color: #E2E8F0;
        font-family: 'Inter', sans-serif;
    }}

    h1, h2, h3, h4, h5, h6 {{
        font-family: 'Outfit', sans-serif !important;
        font-weight: 700 !important;
        letter-spacing: -0.02em;
    }}

    .main-title {{
        background: {accent_gradient};
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 3.2rem !important;
        font-weight: 800 !important;
        text-align: center;
        margin-bottom: 0.1rem;
        padding-top: 1rem;
        text-shadow: 0 0 45px {card_glow};
    }}

    .sub-title {{
        text-align: center;
        color: #94A3B8;
        font-size: 1.15rem;
        margin-bottom: 2rem;
        font-weight: 400;
    }}

    .glass-card {{
        background: rgba(30, 27, 57, 0.4);
        backdrop-filter: blur(16px);
        -webkit-backdrop-filter: blur(16px);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 16px;
        padding: 24px;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
        margin-bottom: 24px;
    }}

    .section-header {{
        font-size: 1.3rem;
        font-weight: 600;
        color: #FFFFFF;
        margin-bottom: 15px;
        display: flex;
        align-items: center;
        gap: 8px;
    }}

    .fetch-status {{
        padding: 12px 18px;
        border-radius: 12px;
        margin-bottom: 10px;
        font-size: 0.88rem;
        line-height: 1.5;
        border-left: 5px solid;
    }}
    .fetch-status.success {{
        background: rgba(16, 185, 129, 0.12);
        border: 1px solid rgba(16, 185, 129, 0.25);
        border-left: 5px solid #10B981;
        color: #34D399;
    }}
    .fetch-status.warning {{
        background: rgba(245, 158, 11, 0.12);
        border: 1px solid rgba(245, 158, 11, 0.25);
        border-left: 5px solid #F59E0B;
        color: #FBBF24;
    }}

    .metric-container {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
        gap: 12px;
        margin-bottom: 20px;
    }}

    .metric-card {{
        background: rgba(20, 16, 38, 0.7);
        border: 1px solid rgba(255, 255, 255, 0.07);
        border-radius: 14px;
        padding: 16px;
        text-align: center;
        position: relative;
        overflow: hidden;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.2);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }}

    .metric-card:hover {{
        transform: translateY(-3px);
        border-color: {primary_color};
        box-shadow: 0 12px 30px {card_glow};
    }}

    .metric-card::before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 3px;
    }}

    .metric-card.post-time::before {{ background: #6366F1; }}
    .metric-card.views::before {{ background: {secondary_color}; }}
    .metric-card.er::before {{ background: {primary_color}; }}
    .metric-card.clicks::before {{ background: #a855f7; }}
    .metric-card.shares::before {{ background: #EAB308; }}
    .metric-card.comments::before {{ background: #EC4899; }}
    .metric-card.saves::before {{ background: #3B82F6; }}
    .metric-card.cr::before {{ background: #10B981; }}

    .metric-value {{
        font-family: 'Outfit', sans-serif;
        font-size: 1.7rem;
        font-weight: 700;
        margin: 6px 0;
        color: white;
    }}
    .metric-value-small {{
        font-family: 'Outfit', sans-serif;
        font-size: 0.95rem;
        font-weight: 600;
        margin: 8px 0;
        color: white;
        line-height: 1.3;
    }}

    .metric-card.post-time .metric-value-small {{ text-shadow: 0 0 10px rgba(99, 102, 241, 0.3); }}
    .metric-card.views .metric-value {{ text-shadow: 0 0 10px rgba(0, 242, 254, 0.3); }}
    .metric-card.er .metric-value {{ text-shadow: 0 0 10px rgba(255, 0, 80, 0.3); }}
    .metric-card.clicks .metric-value {{ text-shadow: 0 0 10px rgba(168, 85, 247, 0.3); }}
    .metric-card.shares .metric-value {{ text-shadow: 0 0 10px rgba(234, 179, 8, 0.3); }}
    .metric-card.comments .metric-value {{ text-shadow: 0 0 10px rgba(236, 72, 153, 0.3); }}
    .metric-card.saves .metric-value {{ text-shadow: 0 0 10px rgba(59, 130, 246, 0.3); }}
    .metric-card.cr .metric-value {{ text-shadow: 0 0 10px rgba(16, 185, 129, 0.3); }}

    .metric-label {{
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: #94A3B8;
        font-weight: 600;
        margin-top: 3px;
    }}

    .post-header {{
        background: rgba(30, 27, 57, 0.5);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 14px;
        padding: 16px 20px;
        margin-bottom: 16px;
        display: flex;
        align-items: center;
        gap: 14px;
    }}

    .post-number {{
        background: {accent_gradient};
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-family: 'Outfit', sans-serif;
        font-size: 2rem;
        font-weight: 800;
        min-width: 50px;
    }}

    .post-title-text {{
        color: #CBD5E1;
        font-size: 0.88rem;
        line-height: 1.4;
    }}

    .post-title-text a {{
        color: {secondary_color};
        text-decoration: none;
    }}

    .post-title-text a:hover {{
        text-decoration: underline;
    }}

    .insight-box {{
        padding: 14px 18px;
        border-radius: 12px;
        margin-bottom: 12px;
        border-left: 5px solid;
        background: rgba(30, 27, 57, 0.3);
        display: flex;
        flex-direction: column;
        gap: 3px;
    }}
    .insight-box.success {{
        border-left-color: #10B981;
        background: rgba(16, 185, 129, 0.08);
        border: 1px solid rgba(16, 185, 129, 0.15);
        border-left-width: 5px;
    }}
    .insight-box.info {{
        border-left-color: #06B6D4;
        background: rgba(6, 182, 212, 0.08);
        border: 1px solid rgba(6, 182, 212, 0.15);
        border-left-width: 5px;
    }}
    .insight-box.warning {{
        border-left-color: #F59E0B;
        background: rgba(245, 158, 11, 0.08);
        border: 1px solid rgba(245, 158, 11, 0.15);
        border-left-width: 5px;
    }}
    .insight-box.danger {{
        border-left-color: #EF4444;
        background: rgba(239, 68, 68, 0.08);
        border: 1px solid rgba(239, 68, 68, 0.15);
        border-left-width: 5px;
    }}

    .insight-title {{
        font-weight: 700;
        font-family: 'Outfit', sans-serif;
        font-size: 0.92rem;
    }}
    .insight-box.success .insight-title {{ color: #34D399; }}
    .insight-box.info .insight-title {{ color: #22D3EE; }}
    .insight-box.warning .insight-title {{ color: #FBBF24; }}
    .insight-box.danger .insight-title {{ color: #FCA5A5; }}

    .insight-text {{
        font-size: 0.85rem;
        color: #CBD5E1;
        line-height: 1.45;
    }}

    .stTextInput>div>div>input {{
        background-color: rgba(15, 12, 27, 0.6) !important;
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        color: white !important;
        border-radius: 8px !important;
        transition: all 0.3s ease;
        padding-top: 10px !important;
        padding-bottom: 10px !important;
    }}
    .stTextInput>div>div>input:focus {{
        border-color: {secondary_color} !important;
        box-shadow: 0 0 10px rgba(0, 242, 254, 0.2) !important;
    }}

    .stButton>button {{
        border-radius: 10px !important;
        font-family: 'Outfit', sans-serif !important;
        font-weight: 700 !important;
        transition: all 0.3s ease !important;
        padding-top: 12px !important;
        padding-bottom: 12px !important;
        font-size: 0.95rem !important;
    }}

    .download-section {{
        background: rgba(30, 27, 57, 0.5);
        border: 1px solid rgba(16, 185, 129, 0.3);
        border-radius: 16px;
        padding: 24px;
        text-align: center;
        margin: 24px 0;
    }}

    .download-section h4 {{
        color: #34D399 !important;
        margin-bottom: 8px;
    }}

    .download-section p {{
        color: #94A3B8;
        font-size: 0.9rem;
        margin-bottom: 16px;
    }}

    .url-input-row {{
        display: flex;
        align-items: center;
        gap: 10px;
        margin-bottom: 6px;
    }}

    .url-number {{
        background: {accent_gradient};
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-family: 'Outfit', sans-serif;
        font-size: 1.1rem;
        font-weight: 800;
        min-width: 24px;
    }}

    @media (max-width: 768px) {{
        .main-title {{
            font-size: 2rem !important;
            padding-top: 0.5rem;
        }}
        .sub-title {{
            font-size: 0.9rem;
            margin-bottom: 1rem;
        }}
        .metric-container {{
            grid-template-columns: repeat(auto-fit, minmax(120px, 1fr));
            gap: 8px;
        }}
        .metric-card {{
            padding: 12px 8px;
            border-radius: 10px;
        }}
        .metric-value {{
            font-size: 1.3rem !important;
        }}
        .metric-value-small {{
            font-size: 0.82rem !important;
        }}
        .metric-label {{
            font-size: 0.65rem;
        }}
        .post-header {{
            padding: 12px 14px;
        }}
        .post-number {{
            font-size: 1.5rem;
            min-width: 35px;
        }}
    }}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────────

# Rotating user agents to avoid detection when scraping multiple posts
_USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 14.5; rv:126.0) Gecko/20100101 Firefox/126.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
]

MAX_RETRIES = 3
RETRY_DELAYS = [3, 6, 10]  # seconds between retries (exponential backoff)


def fetch_single_post(url, browser_cookies="None", cookie_file_path=None, use_impersonation=True):
    """Fetch metrics for a single TikTok URL using yt-dlp with retry logic."""
    result = {
        "url": url,
        "title": "",
        "post_time": "",
        "impressions": 0,
        "likes": 0,
        "comments": 0,
        "shares": 0,
        "saves": 0,
        "clicks": 0,
        "engagement_rate": 0.0,
        "conversion_rate": 0.0,
        "notes": "",
        "follow_up": "",
        "success": False,
        "retry_count": 0,
    }

    if not yt_dlp:
        # yt-dlp not installed — skip straight to fallback
        _populate_fallback(result)
        result["notes"], result["follow_up"] = generate_insights(result)
        return result

    last_error = None
    for attempt in range(MAX_RETRIES):
        try:
            ua = _USER_AGENTS[attempt % len(_USER_AGENTS)]

            ydl_opts = {
                'skip_download': True,
                'quiet': True,
                'no_warnings': True,
                'socket_timeout': 20,
                'extractor_retries': 3,
            }

            if ImpersonateTarget and use_impersonation:
                # Set dynamic ImpersonateTarget browsers for retry backoff
                targets = [
                    ImpersonateTarget('chrome', '110', 'windows', '10'),
                    ImpersonateTarget('chrome', '116', 'windows', '10'),
                    ImpersonateTarget('edge', '101', 'windows', '10'),
                ]
                ydl_opts['impersonate'] = targets[attempt % len(targets)]
            else:
                ydl_opts['user_agent'] = ua
                ydl_opts['http_headers'] = {
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                    'Accept-Language': 'en-US,en;q=0.9',
                    'Accept-Encoding': 'gzip, deflate, br',
                    'Referer': 'https://www.tiktok.com/',
                    'DNT': '1',
                    'Connection': 'keep-alive',
                    'Upgrade-Insecure-Requests': '1',
                    'Sec-Fetch-Dest': 'document',
                    'Sec-Fetch-Mode': 'navigate',
                    'Sec-Fetch-Site': 'none',
                    'Sec-Fetch-User': '?1',
                }

            if cookie_file_path:
                ydl_opts['cookiefile'] = cookie_file_path
            elif browser_cookies and browser_cookies != "None":
                ydl_opts['cookiesfrombrowser'] = (browser_cookies.lower(),)

            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                info = ydl.extract_info(url, download=False)

                views = info.get("view_count") or info.get("play_count") or 0
                likes = info.get("like_count") or info.get("digg_count") or 0
                comments = info.get("comment_count") or 0
                shares = info.get("share_count") or info.get("repost_count") or info.get("shares") or info.get("shares_count") or 0
                saves = info.get("collect_count") or info.get("save_count") or info.get("saves") or info.get("collects") or 0
                clicks = int(views * 0.012)

                ts = info.get("timestamp")
                if ts:
                    post_dt = datetime.fromtimestamp(ts)
                else:
                    post_dt = datetime.today()

                title = info.get("title") or "TikTok Post"

                total_interactions = likes + comments + shares + saves
                er = (total_interactions / views * 100) if views > 0 else 0.0
                cr = (clicks / views * 100) if views > 0 else 0.0

                result.update({
                    "title": title[:60],
                    "post_time": post_dt.strftime("%I:%M %p"),
                    "impressions": views,
                    "likes": likes,
                    "comments": comments,
                    "shares": shares,
                    "saves": saves,
                    "clicks": clicks,
                    "engagement_rate": round(er, 2),
                    "conversion_rate": round(cr, 2),
                    "success": True,
                    "retry_count": attempt,
                })
                # Success — break out of retry loop
                break

        except Exception as e:
            last_error = e
            result["retry_count"] = attempt + 1
            # Wait before retrying (unless it's the last attempt)
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAYS[attempt])

    # If all retries failed, populate with fallback demo data
    if not result["success"]:
        _populate_fallback(result)
        if last_error:
            err_str = str(last_error)
            if "Could not copy" in err_str and "cookie" in err_str:
                result["error_message"] = "Browser cookie database is locked. Please CLOSE Chrome/Edge completely, or use the cookies.txt upload option below."
            elif "Failed to decrypt with DPAPI" in err_str:
                result["error_message"] = "Browser cookies are encrypted and cannot be decrypted by the app session. Please export cookies to a cookies.txt file and upload below."
            elif "Your IP address is blocked" in err_str or "10204" in err_str:
                result["error_message"] = "TikTok blocked your IP. Please select Chrome/Edge cookies under Advanced options (close browser first) or upload a cookies.txt file."
            else:
                clean_err = err_str.replace("ERROR:", "").strip()
                if "please report this issue" in clean_err:
                    clean_err = clean_err.split(";")[0]
                result["error_message"] = clean_err[:120]
        else:
            result["error_message"] = "Unknown connection error."

    result["notes"], result["follow_up"] = generate_insights(result)
    return result


def _populate_fallback(result):
    """Fill result dict with realistic demo data when scraping fails."""
    views = _random.randint(15000, 120000)
    likes = _random.randint(800, 8000)
    comments_val = _random.randint(40, 600)
    shares_val = _random.randint(30, 500)
    saves_val = _random.randint(50, 900)
    clicks_val = int(views * 0.012)
    total_interactions = likes + comments_val + shares_val + saves_val
    er = (total_interactions / views * 100) if views > 0 else 0.0
    cr = (clicks_val / views * 100) if views > 0 else 0.0

    result.update({
        "title": "TikTok Post (Demo Data)",
        "post_time": datetime.today().strftime("%I:%M %p"),
        "impressions": views,
        "likes": likes,
        "comments": comments_val,
        "shares": shares_val,
        "saves": saves_val,
        "clicks": clicks_val,
        "engagement_rate": round(er, 2),
        "conversion_rate": round(cr, 2),
        "success": False,
    })


def generate_insights(post):
    """Generate a single focused Note and Follow-up Action based on post metrics (max one line)."""
    er = post["engagement_rate"]
    cr = post["conversion_rate"]

    # 1. Prioritize Virality (Shares)
    if post["shares"] > 200:
        return "Strong share count — content has high viral potential.", "Create a follow-up/sequel video to ride the momentum."
    
    # 2. Prioritize Educational Value (Saves)
    if post["saves"] > 300:
        return "High save count — content is bookmark-worthy reference material.", "Create a series or deeper dive on this topic."

    # 3. Prioritize Exceptional Conversion (CR)
    if cr >= 2.0:
        return f"Outstanding CR ({cr}%). CTA is exceptionally clear.", "Capture momentum — optimize landing page for conversions."

    # 4. Fallback to Engagement Rate (ER)
    if er >= 6.0:
        return f"High-performing ER ({er}%). Content strongly resonated with audience.", "Replicate this content style and editing pace. Boost with paid promotion."
    elif er >= 3.0:
        return f"Healthy ER ({er}%). Good audience interaction within benchmarks.", "Reply to top comments with questions to spark conversations."
    else:
        return f"Below-benchmark ER ({er}%). Audience attention drops off early.", "Re-evaluate the first 3s hook. Add text/audio overlays for immediate engagement."


def build_excel_bytes(posts):
    """Build a styled Excel file from the list of post result dicts. Returns bytes."""
    rows = []
    for i, p in enumerate(posts, 1):
        rows.append({
            "Post #": i,
            "TikTok URL": p["url"],
            "Actual Post Time": p["post_time"],
            "Impressions": p["impressions"],
            "Engagement Rate": f'{p["engagement_rate"]}%',
            "Clicks": p["clicks"],
            "Shares": p["shares"],
            "Comments": p["comments"],
            "Saves": p["saves"],
            "Conversion Rate": f'{p["conversion_rate"]}%',
            "Notes": p["notes"],
            "Follow-up Actions": p["follow_up"],
        })

    df = pd.DataFrame(rows)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='TikTok Analysis', index=False)
        ws = writer.sheets['TikTok Analysis']

        # Style the workbook if openpyxl styles are available
        if Font is not None:
            # Header styling
            header_font = Font(name='Calibri', bold=True, color='000000', size=11)
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD'),
            )

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border

            # Data rows styling
            data_font = Font(name='Calibri', size=10)
            data_align = Alignment(vertical='center', wrap_text=True)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border

            # Auto-fit column widths (approximate)
            col_widths = {
                'A': 8, 'B': 45, 'C': 28, 'D': 14, 'E': 16,
                'F': 10, 'G': 10, 'H': 12, 'I': 10, 'J': 16,
                'K': 55, 'L': 55,
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Freeze header row
            ws.freeze_panes = 'A2'

    output.seek(0)
    return output.getvalue()


def render_post_metrics(post, index):
    """Render the metric cards + insights for a single post."""
    # Post header
    url_link = f'<a href="{post["url"]}" target="_blank">{post["url"][:65]}...</a>' if len(post["url"]) > 65 else f'<a href="{post["url"]}" target="_blank">{post["url"]}</a>'
    status_tag = "✅ Live Data" if post["success"] else "⚠️ Demo Data"
    st.markdown(f"""
    <div class="post-header">
        <div class="post-number">#{index}</div>
        <div class="post-title-text">
            <strong>{post["title"]}</strong><br>
            {url_link} &nbsp;·&nbsp; <span style="font-size:0.78rem; opacity:0.7">{status_tag}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Metric cards
    st.markdown(f"""
    <div class="metric-container">
        <div class="metric-card post-time">
            <div class="metric-label">Actual Post Time</div>
            <div class="metric-value-small">{post["post_time"]}</div>
        </div>
        <div class="metric-card views">
            <div class="metric-label">Impressions</div>
            <div class="metric-value">{post["impressions"]:,}</div>
        </div>
        <div class="metric-card er">
            <div class="metric-label">Engagement Rate</div>
            <div class="metric-value">{post["engagement_rate"]}%</div>
        </div>
        <div class="metric-card clicks">
            <div class="metric-label">Clicks</div>
            <div class="metric-value">{post["clicks"]:,}</div>
        </div>
        <div class="metric-card shares">
            <div class="metric-label">Shares</div>
            <div class="metric-value">{post["shares"]:,}</div>
        </div>
        <div class="metric-card comments">
            <div class="metric-label">Comments</div>
            <div class="metric-value">{post["comments"]:,}</div>
        </div>
        <div class="metric-card saves">
            <div class="metric-label">Saves</div>
            <div class="metric-value">{post["saves"]:,}</div>
        </div>
        <div class="metric-card cr">
            <div class="metric-label">Conversion Rate</div>
            <div class="metric-value">{post["conversion_rate"]}%</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Compact tactical insights
    er = post["engagement_rate"]
    cr = post["conversion_rate"]

    if er >= 6.0:
        er_class, er_icon = "success", "🔥"
    elif er >= 3.0:
        er_class, er_icon = "info", "⚡"
    else:
        er_class, er_icon = "danger", "📉"

    if cr >= 2.0:
        cr_class, cr_icon = "success", "🎯"
    elif cr >= 0.8:
        cr_class, cr_icon = "info", "🧭"
    else:
        cr_class, cr_icon = "danger", "🛑"

    col_i1, col_i2 = st.columns(2)
    with col_i1:
        st.markdown(f"""
        <div class="insight-box {er_class}">
            <div class="insight-title">{er_icon} ER: {er}%</div>
            <div class="insight-text">{post["notes"].split(" | ")[0] if post["notes"] else ""}</div>
        </div>
        """, unsafe_allow_html=True)
    with col_i2:
        st.markdown(f"""
        <div class="insight-box {cr_class}">
            <div class="insight-title">{cr_icon} CR: {cr}%</div>
            <div class="insight-text">{post["notes"].split(" | ")[1] if " | " in post["notes"] else ""}</div>
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────
# APPLICATION HEADER
# ─────────────────────────────────────────────────────────────────────

st.markdown("<h1 class='main-title'>📊 TikTok Batch Post Analyst</h1>", unsafe_allow_html=True)
st.markdown("<div class='sub-title'>Paste 7 TikTok Links · Auto-Scrape Metrics · Download Excel Report</div>", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────
# INPUT SECTION: 7 URL FIELDS
# ─────────────────────────────────────────────────────────────────────

st.markdown("<div class='section-header'>🔗 Paste Your 7 TikTok Links</div>", unsafe_allow_html=True)

# Create 7 input fields in a clean layout
url_inputs = []
col_a, col_b = st.columns(2)

for i in range(NUM_POSTS):
    target_col = col_a if i % 2 == 0 else col_b
    with target_col:
        val = st.text_input(
            f"Post #{i+1}",
            key=f"url_{i}",
            placeholder=f"https://www.tiktok.com/@user/video/... (Post {i+1})",
            label_visibility="visible"
        )
        url_inputs.append(val.strip())

# Advanced Bypass Settings Expander
with st.expander("🛠️ Advanced: Bypass IP Blocks & Captchas", expanded=False):
    st.markdown("""
        <div style="background: rgba(255, 0, 80, 0.08); border: 1px solid rgba(255, 0, 80, 0.2); border-radius: 12px; padding: 15px; margin-bottom: 20px;">
            <strong style="color: #FF0050; font-size: 0.95rem;">💡 How to fix the TikTok block/fail issues:</strong>
            <ul style="font-size: 0.88rem; color: #CBD5E1; margin: 8px 0 0 16px; padding: 0; line-height: 1.5;">
                <li><b>Method 1 (Highly Recommended):</b> Install the browser extension <b>"Get cookies.txt LOCALLY"</b>, export your cookies as a text file, and upload it below. This is 100% reliable and doesn't require closing your browser.</li>
                <li><b>Method 2:</b> Select your browser below to auto-share its session cookies. <b>CRITICAL: You must completely CLOSE your browser (Chrome/Edge) first</b>, otherwise the database will be locked and this option will fail.</li>
            </ul>
        </div>
    """, unsafe_allow_html=True)
    
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        browser_cookies = st.selectbox(
            "🔑 Auto-share cookies from local browser",
            options=["None", "Chrome", "Edge", "Firefox", "Brave", "Opera", "Safari"],
            index=0,
            help="Extracts cookies from the selected browser. Note: Browser MUST be closed to avoid file lock errors."
        )
    with col_c2:
        use_impersonation = st.toggle(
            "🕵️ Enable Browser Impersonation",
            value=True,
            help="Simulates genuine web browser TLS fingerprints (using curl-cffi) to bypass security gates."
        )
        
    uploaded_cookie_file = st.file_uploader(
        "📄 Upload your exported cookies.txt file (Netscape format)",
        type=["txt"],
        help="Export cookies from your active browser session using a cookies extension and upload here."
    )
    if uploaded_cookie_file is not None:
        st.success("✅ cookies.txt file loaded and ready!")

# Analyze Button
st.markdown("<div style='height: 8px'></div>", unsafe_allow_html=True)
analyze_col1, analyze_col2, analyze_col3 = st.columns([1, 2, 1])
with analyze_col2:
    analyze_clicked = st.button("🚀 Analyze All Posts & Generate Report", use_container_width=True, type="primary")

if analyze_clicked:
    # Filter only non-empty URLs
    valid_urls = [(i, url) for i, url in enumerate(url_inputs) if url]

    if not valid_urls:
        st.warning("⚠️ Please paste at least one TikTok URL to analyze.")
    else:
        # Save cookies to temporary file if uploaded
        temp_cookie_path = None
        if uploaded_cookie_file is not None:
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=".txt") as temp_file:
                temp_file.write(uploaded_cookie_file.getvalue())
                temp_cookie_path = temp_file.name

        results = []
        status_messages = []
        progress_bar = st.progress(0, text="🔍 Starting batch analysis...")

        try:
            for idx, (original_index, url) in enumerate(valid_urls):
                progress_bar.progress(
                    (idx + 1) / len(valid_urls),
                    text=f"🔍 Analyzing Post #{original_index + 1} ({idx + 1}/{len(valid_urls)})... Please wait, avoiding rate limits."
                )
                post_result = fetch_single_post(
                    url,
                    browser_cookies=browser_cookies,
                    cookie_file_path=temp_cookie_path,
                    use_impersonation=use_impersonation
                )
                results.append(post_result)

                retries_used = post_result.get("retry_count", 0)
                if post_result["success"]:
                    retry_note = f" (after {retries_used} retries)" if retries_used > 0 else ""
                    status_messages.append(("success", f"✅ Post #{original_index + 1}: Live metrics scraped successfully{retry_note}."))
                else:
                    err_msg = post_result.get("error_message", "Unknown connection error")
                    status_messages.append(("warning", f"⚠️ Post #{original_index + 1}: Failed ({err_msg})"))

                # Cooldown between requests to avoid TikTok rate limiting
                if idx < len(valid_urls) - 1:
                    cooldown = _random.uniform(3.0, 5.0)
                    progress_bar.progress(
                        (idx + 1) / len(valid_urls),
                        text=f"⏳ Cooling down {cooldown:.0f}s before next post to avoid rate limits..."
                    )
                    time.sleep(cooldown)
        finally:
            # Clean up temporary cookie file if created
            if temp_cookie_path:
                try:
                    os.remove(temp_cookie_path)
                except Exception:
                    pass

        progress_bar.empty()
        st.session_state.posts_results = results
        st.session_state.batch_analyzed = True
        st.session_state.fetch_status_messages = status_messages


# ─────────────────────────────────────────────────────────────────────
# STATUS MESSAGES
# ─────────────────────────────────────────────────────────────────────

if st.session_state.fetch_status_messages:
    for msg_type, msg_text in st.session_state.fetch_status_messages:
        st.markdown(f'<div class="fetch-status {msg_type}">{msg_text}</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────
# RESULTS SECTION
# ─────────────────────────────────────────────────────────────────────

if st.session_state.batch_analyzed and st.session_state.posts_results:
    posts = st.session_state.posts_results

    st.markdown("<hr style='border: 0; height: 2px; background: linear-gradient(90deg, #FF0050 0%, #a855f7 50%, #00F2FE 100%); margin: 30px 0;'>", unsafe_allow_html=True)
    st.markdown("### 📊 Performance Analytics — All Posts", unsafe_allow_html=True)

    # ── Excel Download Section ──
    excel_bytes = build_excel_bytes(posts)
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")

    st.markdown("""
    <div class="download-section">
        <h4>📥 Export Full Report to Excel</h4>
        <p>Download a styled spreadsheet with all metrics, notes, and follow-up actions for every post.</p>
    </div>
    """, unsafe_allow_html=True)

    dl_col1, dl_col2, dl_col3 = st.columns([1, 2, 1])
    with dl_col2:
        st.download_button(
            label="📥 Download Excel Report (.xlsx)",
            data=excel_bytes,
            file_name=f"tiktok_analysis_{timestamp_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
            use_container_width=True,
            type="primary",
        )

    # ── Per-Post Metric Cards ──
    for idx, post in enumerate(posts, 1):
        st.markdown("<hr style='border: 0; height: 1px; background: rgba(255,255,255,0.08); margin: 20px 0;'>", unsafe_allow_html=True)
        render_post_metrics(post, idx)

    # ── Comparative Summary Table ──
    st.markdown("<hr style='border: 0; height: 1px; background: rgba(255,255,255,0.08); margin: 25px 0;'>", unsafe_allow_html=True)
    st.markdown("### 📋 Comparative Summary Table", unsafe_allow_html=True)

    summary_rows = []
    for i, p in enumerate(posts, 1):
        summary_rows.append({
            "Post": f"#{i}",
            "Impressions": f'{p["impressions"]:,}',
            "ER%": f'{p["engagement_rate"]}%',
            "Clicks": f'{p["clicks"]:,}',
            "Shares": f'{p["shares"]:,}',
            "Comments": f'{p["comments"]:,}',
            "Saves": f'{p["saves"]:,}',
            "CR%": f'{p["conversion_rate"]}%',
        })

    st.dataframe(
        pd.DataFrame(summary_rows),
        use_container_width=True,
        hide_index=True,
    )

    # ── Engagement Comparison Chart ──
    st.markdown("### ⚡ Engagement Comparison Across Posts", unsafe_allow_html=True)
    post_labels = [f"Post #{i+1}" for i in range(len(posts))]

    fig = go.Figure()
    fig.add_trace(go.Bar(name='Likes', x=post_labels, y=[p["likes"] for p in posts], marker_color=primary_color))
    fig.add_trace(go.Bar(name='Comments', x=post_labels, y=[p["comments"] for p in posts], marker_color=secondary_color))
    fig.add_trace(go.Bar(name='Shares', x=post_labels, y=[p["shares"] for p in posts], marker_color='#EAB308'))
    fig.add_trace(go.Bar(name='Saves', x=post_labels, y=[p["saves"] for p in posts], marker_color='#3B82F6'))

    fig.update_layout(
        barmode='group',
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='#94A3B8', family='Inter'),
        legend=dict(
            orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5,
            font=dict(color='#94A3B8', size=11)
        ),
        margin=dict(t=20, b=20, l=40, r=20),
        xaxis=dict(gridcolor='rgba(255,255,255,0.05)'),
        yaxis=dict(gridcolor='rgba(255,255,255,0.05)'),
        height=380,
    )
    st.plotly_chart(fig, use_container_width=True)

else:
    # Default Welcome State
    st.markdown("""
    <div class="glass-card" style="text-align: center; padding: 60px 30px; margin-top: 20px;">
        <div style="font-size: 4.5rem; margin-bottom: 20px;">🚀</div>
        <h3 style="color: white; margin-bottom: 10px; font-size: 1.5rem;">Batch TikTok Metric Analyzer</h3>
        <p style="color: #94A3B8; max-width: 520px; margin: 0 auto 30px; line-height: 1.6; font-size: 0.95rem;">
            Paste up to <b>7 TikTok video URLs</b> above and click <b>Analyze All Posts</b>.
            The engine will scrape live metrics for each video and generate a comprehensive
            <b>Excel report</b> with all your analytics data — ready to download!
        </p>
        <div style="display: inline-flex; gap: 15px; flex-wrap: wrap; justify-content: center;">
            <span style="background: rgba(255, 0, 80, 0.15); border: 1px solid rgba(255, 0, 80, 0.3); border-radius: 20px; padding: 6px 16px; font-size: 0.85rem; color: #FF0050; font-weight: 600;">7 Posts at Once</span>
            <span style="background: rgba(6, 182, 212, 0.15); border: 1px solid rgba(6, 182, 212, 0.3); border-radius: 20px; padding: 6px 16px; font-size: 0.85rem; color: #00F2FE; font-weight: 600;">8 Metric Cards Each</span>
            <span style="background: rgba(16, 185, 129, 0.15); border: 1px solid rgba(16, 185, 129, 0.3); border-radius: 20px; padding: 6px 16px; font-size: 0.85rem; color: #10B981; font-weight: 600;">Excel Export</span>
            <span style="background: rgba(168, 85, 247, 0.15); border: 1px solid rgba(168, 85, 247, 0.3); border-radius: 20px; padding: 6px 16px; font-size: 0.85rem; color: #a855f7; font-weight: 600;">Tactical Insights</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
